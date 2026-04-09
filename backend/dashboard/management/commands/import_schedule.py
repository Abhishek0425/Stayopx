"""
import_schedule.py — Reads EXACT day-wise menus from Excel and stores them.

Run from backend/:
    python manage.py import_schedule
"""
import os
from django.core.management.base import BaseCommand
from django.conf import settings


DAYS = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']

SKIP = {
    '', None, '—', '-', 'Fruits', 'BBJ', 'Tea/Coffee', 'Tea/ Coffee', 'Tea',
    'Pickle', 'Fryums', 'Green Salad', 'Raita', 'Hot Milk', 'Ketchup',
    'Steamed Rice', 'Desi Ghee Chapatti', ' Chapatti', 'Chapatti',
    'Green Chutney', 'Chutney', ' Chutney', 'Curd/Pickle', 'Curd',
    'Bhajji', 'Dry Aloo', 'Bhajji/ Chutney', 'Curd/Pickle/Chutney',
    'Curd/Green Chutney', 'Curd/Pickle/ bhaji', 'Aloo Bhaji', 'Aloo Rassa',
    'Aloo Jhol', 'Missal/ Curd', 'Missal', 'Matar', 'Mattara',
    'Weekly Menu', 'Day', '2', 'Note-Menu is subject to change as per the avaiability of seasonal vegetabes.',
    'Khata/Mittha Pani', 'Khata/Mittha Pani',
    'Jeera Pulao', 'Corn Pulao', 'Pea Pulao', 'Dum Pulao', 'Plain Rice',
    'Kathal Biryani', 'Navratan Rice', 'Curd rice', 'Veg Biryani',
    'Mix Veg Pulao', 'Fried Rice', 'Herb Rice with Ratatouille',
    'Garlic Bread', 'Pav', 'Litti', 'Bhati', 'Bhature', 'Tawa Kulcha',
    'Laccha Parataha', 'Poori', 'Pasta Salad', 'Green Salad',
    'Halwa', 'Shahi Tukda', 'Gulab Jamun', 'Rice Kheer', 'Vermcilli Kheer',
    'Jalebi', 'Churma', 'Nariyal Laddu', 'Besan Laddu', 'Guldana',
    'Fruit Custard', 'Balushahi', 'Hakka Noodles',
    'Salsa Sauce', 'Paapad ki sabji',   # not real dinner veg
}


def _clean(v):
    if v is None: return None
    s = str(v).strip()
    return None if (s in SKIP or not s) else s


def _parse_sheet(ws, brand):
    """Parse one sheet (= one week) into a dict of day_name → meal data."""
    rows = list(ws.iter_rows(values_only=True))
    current_meal = None
    day_data = {d: dict(bf1=None, bf2=None, l_dal=None, l_veg1=None, l_veg2=None,
                         snack=None, d_dal=None, d_veg1=None, d_veg2=None)
                for d in DAYS}

    for row in rows:
        label = str(row[0]).strip() if row[0] else ''
        vals  = [_clean(row[i]) for i in range(1, 8)]

        # Section headers
        if label == 'BREAKFAST':      current_meal = 'bf'; continue
        if label == 'LUNCH':          current_meal = 'ln'; continue
        if label == 'EVENING SNACKS': current_meal = 'sn'; continue
        if label in ('Dinner','DINNER'): current_meal = 'dn'; continue
        if label in ('Night Milk','MILK','RICE','BREADS','SALAD',
                     'CURD/RAITA','DESERT','PAPAD/PICKLE',
                     'CHUTNEY/CURD/PICKLE','CHUTNEY','BEVERAGE',
                     'Beverages','Bakery','Fruits'): continue
        if not label: continue  # unlabelled Veg 2 rows handled below

        for di, day in enumerate(DAYS):
            v = vals[di] if di < len(vals) else None
            if not v: continue

            if label in ('HOT FOOD',):
                day_data[day]['bf1'] = v
            elif label in ('HOT FOOD 2', 'HoT FOOD 2'):
                day_data[day]['bf2'] = v
            elif label == 'SNACKS':
                day_data[day]['snack'] = v
            elif label == 'DAL' and current_meal == 'ln':
                day_data[day]['l_dal'] = v
            elif label == 'VEG' and current_meal == 'ln':
                day_data[day]['l_veg1'] = v
            elif label == 'Veg 2' and current_meal == 'ln':
                day_data[day]['l_veg2'] = v
            elif label == 'DAL' and current_meal == 'dn':
                day_data[day]['d_dal'] = v
            elif label == 'VEG' and current_meal == 'dn':
                day_data[day]['d_veg1'] = v
            elif label == 'Veg 2' and current_meal == 'dn':
                day_data[day]['d_veg2'] = v

    return day_data


class Command(BaseCommand):
    help = 'Import exact weekly schedules from Excel into ScheduledMeal table.'

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write('Install openpyxl: pip install openpyxl')
            return

        from dashboard.models import ScheduledMeal

        base = settings.BASE_DIR
        tasks = [
            (os.path.join(base, 'UNILIV_Draft_Menu.xlsx'), 'uniliv'),
            (os.path.join(base, 'HUDDLE_Draft_Menu.xlsx'),  'huddle'),
        ]

        ScheduledMeal.objects.all().delete()
        self.stdout.write('Cleared existing schedule.\n')
        total = 0

        for filepath, brand in tasks:
            if not os.path.exists(filepath):
                self.stdout.write(self.style.ERROR(
                    f'File not found: {filepath}\n'
                    f'Please copy your Excel files to backend/ and rename:\n'
                    f'  Draft_Menu__1_.xlsx     → UNILIV_Draft_Menu.xlsx\n'
                    f'  Huddle_Draft_Menu__1_.xlsx → HUDDLE_Draft_Menu.xlsx\n'
                ))
                continue

            wb = openpyxl.load_workbook(filepath, data_only=True)
            self.stdout.write(f'{brand.upper()}: {len(wb.sheetnames)} sheets found')

            for week_num, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                day_data = _parse_sheet(ws, brand)

                for day_name, d in day_data.items():
                    # Skip days with nothing scheduled
                    if not any(d.values()): continue
                    ScheduledMeal.objects.create(
                        brand=brand, week_number=week_num, day_of_week=day_name,
                        breakfast1=d['bf1'],  breakfast2=d['bf2'],
                        lunch_dal=d['l_dal'], lunch_veg1=d['l_veg1'], lunch_veg2=d['l_veg2'],
                        snack=d['snack'],
                        dinner_dal=d['d_dal'], dinner_veg1=d['d_veg1'], dinner_veg2=d['d_veg2'],
                    )
                    total += 1
                self.stdout.write(
                    self.style.SUCCESS(f'  Week {week_num} ({sheet_name}): 7 days imported'))

        self.stdout.write(self.style.SUCCESS(f'\n✅ {total} day-schedules imported.\n'))
        self.stdout.write('Next step: python manage.py runserver\n')
