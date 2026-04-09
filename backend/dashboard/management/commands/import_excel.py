"""
Management command: import dishes from Excel files.

Usage:
    python manage.py import_excel
    python manage.py import_excel --file UNILIV_Weekly_Menu.xlsx --brand uniliv
    python manage.py import_excel --clear  (clears all dishes first)
"""
import os
from django.core.management.base import BaseCommand
from django.conf import settings


MEAL_MAP = {
    'breakfast': 'Breakfast',
    'lunch':     'Lunch',
    'dinner':    'Dinner',
    'snacks':    'Snacks',
    'snack':     'Snacks',
}

SKIP_VALUES = {'—', '-', 'nan', '', 'none', 'n/a'}
SKIP_NAMES  = {'Tea/Coffee', 'Tea', 'Hot Milk', 'Fruits', 'Ketchup',
               'Tea/ Coffee', 'Pickle', 'Fryums', 'BBJ', 'Raita',
               'Green Salad', 'Chutney', 'Green Chutney', 'Steamed Rice',
               'Chapatti', 'Desi Ghee Chapatti'}


def _is_aloo(name):
    return any(w in name.lower() for w in ['aloo', 'potato'])


def _is_dal(name):
    return any(w in name.lower() for w in ['dal','daal','kadhi','rajma','chana',
                                            'moong','arhar','toor','urad','masoor',
                                            'sambhar','sambar','sambher'])


def _is_star(name):
    return any(w in name.lower() for w in ['paneer','biryani','kofta','kadhai',
                                            'makhani','lababdar','shahi'])


def _clean(val):
    s = str(val).strip()
    return None if s.lower() in SKIP_VALUES else s


def _parse_excel(filepath, brand, verbosity):
    try:
        import openpyxl
    except ImportError:
        raise ImportError('openpyxl is required: pip install openpyxl')

    from openpyxl import load_workbook
    wb   = load_workbook(filepath, data_only=True)
    ws   = wb['Weekly Menu']
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else '' for h in rows[0]]

    # Map header → column index
    col = {}
    for i, h in enumerate(headers):
        hl = h.lower()
        if 'day'  in hl:            col['day']  = i
        if 'meal' in hl:            col['meal'] = i
        if 'dal'  == hl:            col['dal']  = i
        if 'main veg 1' in hl:      col['veg1'] = i
        if 'main veg 2' in hl:      col['veg2'] = i
        if 'rice' in hl:            col['rice'] = i
        if 'bread' in hl:           col['bread']= i
        if 'extra' in hl:           col['extra']= i

    dishes = []
    seen   = set()

    for row in rows[1:]:
        meal_raw = _clean(row[col.get('meal', 1)])
        if not meal_raw:
            continue
        meal_type = MEAL_MAP.get(meal_raw.lower())
        if not meal_type:
            continue

        candidates = []

        # Dal column
        dal_name = _clean(row[col.get('dal', 2)])
        if dal_name and dal_name not in SKIP_NAMES:
            candidates.append((dal_name, True, False))

        # Main veg columns
        for key in ('veg1', 'veg2', 'rice', 'bread'):
            val = _clean(row[col.get(key)]) if key in col else None
            if val and val not in SKIP_NAMES:
                candidates.append((val, False, _is_aloo(val)))

        # Extras — split by |
        extras = _clean(row[col.get('extra')]) if 'extra' in col else None
        if extras:
            for part in extras.split('|'):
                p = part.strip()
                if p and p not in SKIP_NAMES and p.lower() not in SKIP_VALUES:
                    candidates.append((p, False, _is_aloo(p)))

        for name, is_dal_flag, is_aloo_flag in candidates:
            key = (name.lower(), meal_type, brand)
            if key in seen:
                continue
            seen.add(key)
            dishes.append({
                'dish_name':   name,
                'meal_type':   meal_type,
                'brand':       brand,
                'is_dal':      is_dal_flag or _is_dal(name),
                'is_aloo':     is_aloo_flag or _is_aloo(name),
                'is_star':     _is_star(name),
                'ingredients': None,
            })

    if verbosity >= 1:
        print(f'  Parsed {len(dishes)} dishes from {os.path.basename(filepath)}')

    return dishes


class Command(BaseCommand):
    help = 'Import dishes from UNILIV and HUDDLE Excel files into the database.'

    def add_arguments(self, parser):
        parser.add_argument('--file',  type=str, default=None, help='Specific Excel file path')
        parser.add_argument('--brand', type=str, default=None, help='Brand override (uniliv/huddle)')
        parser.add_argument('--clear', action='store_true',    help='Delete all dishes before import')

    def handle(self, *args, **options):
        from dashboard.models import Dish

        if options['clear']:
            count = Dish.objects.all().count()
            Dish.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Cleared {count} existing dishes.'))

        # Build file → brand mapping
        base = os.path.dirname(os.path.abspath(__file__))
        backend_dir = settings.BASE_DIR

        if options['file']:
            tasks = [(options['file'], options['brand'] or 'uniliv')]
        else:
            tasks = [
                (os.path.join(backend_dir, 'UNILIV_Weekly_Menu.xlsx'), 'uniliv'),
                (os.path.join(backend_dir, 'HUDDLE_Weekly_Menu.xlsx'), 'huddle'),
            ]

        total_created = 0
        total_skipped = 0

        for filepath, brand in tasks:
            if not os.path.exists(filepath):
                self.stdout.write(self.style.ERROR(f'File not found: {filepath}'))
                continue

            self.stdout.write(f'\nImporting {brand.upper()} from {os.path.basename(filepath)}...')
            dishes = _parse_excel(filepath, brand, options['verbosity'])

            created = 0
            skipped = 0
            for d in dishes:
                obj, was_created = Dish.objects.get_or_create(
                    dish_name=d['dish_name'],
                    meal_type=d['meal_type'],
                    brand=d['brand'],
                    defaults={
                        'ingredients': d['ingredients'],
                        'is_dal':      d['is_dal'],
                        'is_aloo':     d['is_aloo'],
                        'is_star':     d['is_star'],
                    }
                )
                if was_created:
                    created += 1
                else:
                    skipped += 1

            self.stdout.write(
                self.style.SUCCESS(f'  ✅ {created} created, {skipped} skipped (already exist)')
            )
            total_created += created
            total_skipped += skipped

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'Import complete: {total_created} dishes added, {total_skipped} already existed.'
        ))
        self.stdout.write('')
        self.stdout.write('Next steps:')
        self.stdout.write('  1. Run: python manage.py runserver')
        self.stdout.write('  2. Open admin panel to mark ⭐ Star dishes: http://127.0.0.1:8000/admin')
        self.stdout.write('  3. Open index.html and use the Food Analysis Dashboard')
